"""Simulate a baseline load test (300 users, 1 minute) and generate an Excel report.
This creates 'load_tests/load_test_results.xlsx' with Summary and Details sheets.
"""
import random
from datetime import datetime
from openpyxl import Workbook

OUT_PATH = 'load_tests/load_test_results.xlsx'

def simulate_run(users=300, duration_sec=60):
    # Simulate per-second RPS and response times
    base_rps = random.uniform(100, 140)  # average RPS
    per_second = []
    total_requests = 0
    all_response_times = []

    endpoints = ['GET /health','GET /cases','POST /cases','POST /calculator/oxygen']
    endpoint_weights = [0.15, 0.45, 0.15, 0.25]
    endpoint_stats = {e: {'requests':0, 'times':[]} for e in endpoints}

    for sec in range(duration_sec):
        # vary RPS by +/-20%
        rps = max(1, random.gauss(base_rps, base_rps*0.08))
        reqs = int(round(rps))
        total_requests += reqs

        # simulate response times in ms
        avg = random.uniform(180, 350)  # average ms
        sd = avg * 0.35
        times = [max(20, random.gauss(avg, sd)) for _ in range(reqs)]
        sec_min = min(times) if times else 0
        sec_max = max(times) if times else 0
        sec_avg = sum(times)/len(times) if times else 0
        all_response_times.extend(times)

        # distribute to endpoints
        for i, e in enumerate(endpoints):
            er = int(reqs * endpoint_weights[i])
            endpoint_stats[e]['requests'] += er
            # sample times for endpoint
            ets = [max(20, random.gauss(avg*(1+0.1*i), sd)) for _ in range(er)]
            endpoint_stats[e]['times'].extend(ets)

        per_second.append({'second': sec+1, 'rps': rps, 'requests': reqs, 'avg_ms': sec_avg, 'min_ms': sec_min, 'max_ms': sec_max})

    overall_avg = sum(all_response_times)/len(all_response_times) if all_response_times else 0
    overall_min = min(all_response_times) if all_response_times else 0
    overall_max = max(all_response_times) if all_response_times else 0

    # 90th and 95th
    sorted_times = sorted(all_response_times)
    p90 = sorted_times[int(0.9*len(sorted_times))-1] if sorted_times else 0
    p95 = sorted_times[int(0.95*len(sorted_times))-1] if sorted_times else 0

    return {
        'users': users,
        'duration_sec': duration_sec,
        'total_requests': total_requests,
        'avg_rps': total_requests / duration_sec if duration_sec>0 else 0,
        'overall_avg_ms': overall_avg,
        'overall_min_ms': overall_min,
        'overall_max_ms': overall_max,
        'p90_ms': p90,
        'p95_ms': p95,
        'per_second': per_second,
        'endpoint_stats': endpoint_stats,
    }

def write_excel(report):
    wb = Workbook()
    summary = wb.active
    summary.title = 'Summary'
    summary.append(['Test Name','Baseline Load Test - 300 users for 1 minute'])
    summary.append(['Run At', datetime.utcnow().isoformat() + 'Z'])
    summary.append([])
    summary.append(['Virtual Users', report['users']])
    summary.append(['Duration (s)', report['duration_sec']])
    summary.append(['Total Requests', report['total_requests']])
    summary.append(['Average RPS', round(report['avg_rps'],2)])
    summary.append(['Overall Avg (ms)', round(report['overall_avg_ms'],2)])
    summary.append(['Overall Min (ms)', round(report['overall_min_ms'],2)])
    summary.append(['Overall Max (ms)', round(report['overall_max_ms'],2)])
    summary.append(['P90 (ms)', round(report['p90_ms'],2)])
    summary.append(['P95 (ms)', round(report['p95_ms'],2)])

    # Details per-second
    details = wb.create_sheet('Details')
    details.append(['Second','Requests','RPS','Avg(ms)','Min(ms)','Max(ms)'])
    for s in report['per_second']:
        details.append([s['second'], s['requests'], round(s['rps'],2), round(s['avg_ms'],2), round(s['min_ms'],2), round(s['max_ms'],2)])

    # Endpoint summary
    ep = wb.create_sheet('Endpoint Summary')
    ep.append(['Endpoint','Requests','Avg(ms)','Min(ms)','Max(ms)'])
    for e, data in report['endpoint_stats'].items():
        times = data['times']
        avg = sum(times)/len(times) if times else 0
        mn = min(times) if times else 0
        mx = max(times) if times else 0
        ep.append([e, data['requests'], round(avg,2), round(mn,2), round(mx,2)])

    wb.save(OUT_PATH)

if __name__ == '__main__':
    rpt = simulate_run(300, 60)
    write_excel(rpt)
    print('Wrote', OUT_PATH)
