import os
import pytest
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options

TESTCASES_PATH = os.path.join('lib', 'frontend', 'selenium-tests', 'testcases.xlsx')

def load_testcases(path=TESTCASES_PATH):
    wb = load_workbook(path, read_only=True)
    ws = wb['Details']
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    cases = []
    for r in rows:
        # columns: TestID, Title, Description, Steps, Expected, Status, Priority, Module
        tid, title, desc, steps, expected, status, priority, module = r
        cases.append({'id': tid, 'title': title, 'desc': desc, 'steps': steps, 'expected': expected, 'status': status, 'priority': priority, 'module': module})
    return cases

@pytest.fixture(scope='session')
def driver():
    chrome_options = Options()
    chrome_options.add_argument('--headless=new')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-gpu')

    # Expect user has chromedriver on PATH or provide CHROMEDRIVER_PATH env var
    chromedriver = os.environ.get('CHROMEDRIVER_PATH')
    if chromedriver:
        service = ChromeService(executable_path=chromedriver)
    else:
        service = ChromeService()

    drv = webdriver.Chrome(service=service, options=chrome_options)
    yield drv
    drv.quit()


cases = load_testcases()

@pytest.mark.parametrize('case', cases)
def test_frontend_case(driver, case):
    url = os.environ.get('FRONTEND_URL', 'http://localhost:8080')
    driver.get(url)
    title = driver.title or ''
    # Basic smoke assertion: page title must exist
    assert len(title) > 0, f"Page title empty for {case['id']} - {case['title']}"
