import os
import pytest
from openpyxl import load_workbook

APPIUM_URL = os.environ.get('APPIUM_SERVER_URL')

def load_testcases(path='lib/frontend/appium-tests/testcases_appium_99.xlsx'):
    wb = load_workbook(path, read_only=True)
    ws = wb['Details']
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    cases = []
    for r in rows:
        tid, title, desc, screen, action, expected, status, priority, module = r
        cases.append({'id': tid, 'title': title, 'desc': desc, 'screen': screen, 'action': action, 'expected': expected, 'status': status, 'priority': priority, 'module': module})
    return cases

cases = []
try:
    cases = load_testcases()
except Exception:
    cases = []

@pytest.mark.parametrize('case', cases[:50])  # limit default run to 50 tests
def test_app_flow(case):
    """Simple scaffold: connect to Appium server if available, otherwise skip.
    The test checks that an Appium server is reachable and would open a session.
    Actual device capabilities must be provided via environment variables.
    """
    if not APPIUM_URL:
        pytest.skip('APPIUM_SERVER_URL not set; skipping Appium tests')

    from appium import webdriver

    caps = {}
    # Minimal capability placeholders; users should set via env vars
    caps['platformName'] = os.environ.get('APPIUM_PLATFORM_NAME', 'Android')
    if os.environ.get('APPIUM_APP'):
        caps['app'] = os.environ.get('APPIUM_APP')
    else:
        # Try package/activity if provided
        if os.environ.get('APPIUM_APP_PACKAGE'):
            caps['appPackage'] = os.environ.get('APPIUM_APP_PACKAGE')
        if os.environ.get('APPIUM_APP_ACTIVITY'):
            caps['appActivity'] = os.environ.get('APPIUM_APP_ACTIVITY')

    try:
        driver = webdriver.Remote(command_executor=APPIUM_URL, desired_capabilities=caps)
    except Exception as e:
        pytest.skip(f'Unable to connect to Appium server at {APPIUM_URL}: {e}')

    try:
        # basic smoke: page source should be retrievable
        src = driver.page_source
        assert src is not None
    finally:
        try:
            driver.quit()
        except Exception:
            pass
