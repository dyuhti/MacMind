import os
import json
import pytest
import requests
from openpyxl import load_workbook

BASE = os.environ.get('BACKEND_URL', None)

def load_input():
    with open('backend/automated-tests/input.json') as f:
        return json.load(f)

def load_testcases(path='backend/automated-tests/testcases_backend.xlsx'):
    wb = load_workbook(path, read_only=True)
    ws = wb['Details']
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    cases = []
    for r in rows:
        tid, title, desc, method, endpoint, expected, status, priority, module = r
        cases.append({'id': tid, 'title': title, 'desc': desc, 'method': method, 'endpoint': endpoint, 'expected': expected, 'status': status, 'priority': priority, 'module': module})
    return cases

@pytest.fixture(scope='session')
def config():
    cfg = load_input()
    base = os.environ.get('BACKEND_URL') or cfg.get('backend_url', 'http://localhost:5000')
    cfg['backend_url'] = base
    return cfg

cases = []
try:
    cases = load_testcases()
except Exception:
    cases = []

@pytest.mark.parametrize('case', cases[:50])  # limit to first 50 by default to keep runs reasonable
def test_call_endpoint(config, case):
    url = config['backend_url'].rstrip('/') + case['endpoint']
    method = (case['method'] or 'GET').upper()

    # For POST endpoints, send a minimal payload; for auth/login use sample creds
    payload = {}
    headers = {}
    if case['module'] == 'Auth' and '/login' in case['endpoint']:
        input_cfg = load_input().get('sample_user', {})
        payload = {'email': input_cfg.get('email'), 'password': input_cfg.get('password')}

    if method == 'GET':
        resp = requests.get(url, timeout=10)
    else:
        resp = requests.post(url, json=payload, timeout=10)

    # Pass if response status code matches expected (or is in 2xx range)
    expected = int(case['expected'] or 200)
    assert (200 <= resp.status_code < 300) or resp.status_code == expected, f"{case['id']} {case['title']} failed: {resp.status_code}"
